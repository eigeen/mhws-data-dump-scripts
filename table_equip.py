import json
import pandas as pd
from openpyxl.styles import Alignment

from library.excel_auto_fit import ExcelAutoFit
from library.item_db import get_global_item_db
from library.text_db import get_global_text_db
from library.utils import (
    is_guid_like,
    minify_nested_serial,
    remove_enum_value,
    reindex_column,
    rare_enum_to_value,
)
from library.rare import apply_fix_rare_colors
from table_skill import dump_skill_common_data
from parse_whistle_tone import ToneParser
from table_general import dump_enum_maker, load_enum_internal, dump_user3_data_general

text_db = get_global_text_db()
item_db = get_global_item_db()

BOTTLE_ENUM_TO_ITEM_ID = {
    "CLOSE": "ITEM_0702",
    "STRONG": "ITEM_0703",
    "PENETRATE": "ITEM_0704",
    "PARALYSE": "ITEM_0705",
    "POISON": "ITEM_0706",
    "SLEEP": "ITEM_0707",
    "BLAST": "ITEM_0708",
    "STAMINA": "ITEM_0709",
}
GUN_SHELL_TYPE_TO_ITEM_ID = {
    "NORMAL": "ITEM_0037",
    "PENETRATE": "ITEM_0040",
    "SHOT_GUN": "ITEM_0043",
    "GRENADE": "ITEM_0046",
    "MORTAR": "ITEM_0049",
    "SLASH": "ITEM_0068",
    "RYUUGEKI": "ITEM_0180",  # 龙击弹
    "FIRE": "ITEM_0052",
    "WATER": "ITEM_0053",
    "ELEC": "ITEM_0054",
    "ICE": "ITEM_0055",
    "DRAGON": "ITEM_0056",
    "POISON": "ITEM_0057",
    "PARALYSE": "ITEM_0059",
    "SLEEP": "ITEM_0061",
    "KIJIN": "ITEM_0439",  # 鬼人弹
    "KOUKA": "ITEM_0440",  # 硬化弹
    "HEAL": "ITEM_0441",
    "STAMINA": "ITEM_0443",
    "CAPTURE": "ITEM_0442",
}
SLASH_AEX_BIN_TYPE_TO_NAME_ID = {
    "POWER": "RefStatus_0006_003_01_01",
    "ELEMENT": "RefStatus_0006_003_01_02",
    "DRAGON": "RefStatus_0006_003_01_03",
    "STAMINA": "RefStatus_0006_003_01_06",
    "PARALYSE": "RefStatus_0006_003_01_05",
    "POISON": "RefStatus_0006_003_01_04",
}
# WHISTLE_HIGH_FREQ_TYPE_TO_NAME_ID = {

# }
# WHISTLE_HIBIKI_TYPE_TO_NAME_ID = {

# }


def get_weapon_types() -> dict[str, int]:
    return {
        "LongSword": 0,
        "ShortSword": 1,
        "TwinSword": 2,
        "Tachi": 3,
        "Hammer": 4,
        "Whistle": 5,
        "Lance": 6,
        "Gunlance": 7,
        "SlashAxe": 8,
        "ChargeAxe": 9,
        "Rod": 10,
        "Bow": 11,
        "HeavyBowgun": 12,
        "LightBowgun": 13,
    }


def get_bow_bottle_name(enum_name: str) -> str | None:
    item_id = BOTTLE_ENUM_TO_ITEM_ID.get(enum_name.upper())
    if not item_id:
        return None
    item_entry = item_db.get_entry_by_id(item_id)
    if not item_entry:
        return None
    return item_entry.raw_name


def get_gun_shell_type_name(enum_name: str) -> str | None:
    item_id = GUN_SHELL_TYPE_TO_ITEM_ID.get(enum_name.upper())
    if not item_id:
        return None
    item_entry = item_db.get_entry_by_id(item_id)
    if not item_entry:
        return None
    return item_entry.raw_name


def get_slash_axe_bin_name(bin_type: str) -> str | None:
    name_id = SLASH_AEX_BIN_TYPE_TO_NAME_ID.get(bin_type.upper())
    if not name_id:
        return None
    name = text_db.get_text_by_name(name_id)
    return name


# 将弓箭瓶子list转换为可用的瓶子名字列表
def process_loading_bin(
    loading_bin: list[bool], enum_internal: dict[str, dict]
) -> list[str]:
    if len(loading_bin) != 8:
        raise ValueError("Loading bin length should be 8")
    # 弓箭瓶子处理
    bottle_type_enum = enum_internal["app.Wp11Def.BOTTLE_TYPE"]
    # 反转字典并-1，int->str
    bottle_type_enum_inv = {(v - 1): k for k, v in bottle_type_enum.items()}
    bottle_names = []
    for idx, bottle_type_on in enumerate(loading_bin):
        if not bottle_type_on:
            continue
        bottle_type = bottle_type_enum_inv.get(idx)
        if bottle_type is None:
            print(f"Unknown bottle type: {bottle_type_on}")
            bottle_type = "Unknown"
        bottle_name = get_bow_bottle_name(bottle_type)
        bottle_names.append(bottle_name)

    return bottle_names


# 将弩炮弹药等级表转换为可用的弹药名字列表
def process_gun_shell(
    shell_lv_list: list[str],
    enum_internal: dict[str, dict],
    keep_none_level: bool = False,
) -> list[str]:
    if len(shell_lv_list) != 20:
        raise ValueError("Shell active list length should be 20")
    shell_type_enum = enum_internal["app.WeaponGunDef.SHELL_TYPE"]
    # 反转字典，int->str
    shell_type_enum_inv = {v: k for k, v in shell_type_enum.items()}
    shell_names = []
    for idx, shell_lv in enumerate(shell_lv_list):
        if not keep_none_level and shell_lv == "NONE":
            shell_names.append(None)
            continue
        shell_type = shell_type_enum_inv.get(idx)
        if shell_type is None:
            print(f"Unknown shell type: {shell_lv}")
            shell_names.append(None)
            continue
        shell_name = get_gun_shell_type_name(shell_type)
        shell_names.append(shell_name)

    return shell_names


def dump_armor_series_enum_maker() -> pd.DataFrame:
    path = "natives/STM/GameDesign/Player/EnumMaker/ArmorSeries.user.3.json"
    return dump_enum_maker(path)


def _dump_weapon_data(
    path: str,
    weapon_type: str,
    weapon_types: dict[str, int],
    skill_common_data: pd.DataFrame,
    enum_internal: dict[str, dict],
    whistle_high_freq_data: pd.DataFrame,
    whistle_hibiki_data: pd.DataFrame,
    keep_serial_id: bool = False,
) -> pd.DataFrame:
    data = None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    table = []
    for cData in data[0]["app.user_data.WeaponData"]["_Values"]:
        row = {}
        for key, value in cData["app.user_data.WeaponData.cData"].items():
            if key.startswith("_"):
                key = key[1:]
            key_lower = key.lower()
            if key_lower == weapon_type:
                key = "Id"
            if key_lower in weapon_types and key_lower != weapon_type:
                continue
            # 批量处理列名带 wpxx 的列，排除无关列
            exclude = False
            for wp_type, wp_type_id in weapon_types.items():
                wp_short_id = f"wp{wp_type_id:02d}"
                if key_lower.find(wp_short_id) != -1 and wp_type != weapon_type:
                    exclude = True
                    break
            if exclude:
                continue
            if weapon_type != "rod" and key == "RodInsectLv":
                continue
            if weapon_type not in {"heavybowgun", "lightbowgun"} and key in {
                "MainShell",
                "ShellLv",
                "ShellNum",
                "CustomizePattern",
                "DispSilencer",
                "DispBarrel",
            }:
                continue
            if weapon_type != "lightbowgun" and key in {"RapidShellNum", "IsRappid"}:
                continue
            if weapon_type != "heavybowgun" and key in {
                "EnergyEfficiency",
                "AmmoStrength",
                "EnergyShellTypeNormal",
                "EnergyShellTypeNormal",
                "EnergyShellTypePower",
                "EnergyShellTypeWeak",
            }:
                continue
            if weapon_type != "bow" and key == "isLoadingBin":
                continue
            if weapon_type in {"lightbowgun", "heavybowgun", "bow"} and key in {
                "SharpnessValList",
                "TakumiValList",
            }:
                continue

            value = minify_nested_serial(value)
            if not keep_serial_id:
                value = remove_enum_value(value)

            if key == "Skill":
                # 处理Skill
                for i, skill in enumerate(value):
                    if skill.find("NONE") != -1:
                        continue
                    skill_name = skill_common_data.loc[
                        skill_common_data["skillId"] == skill, "skillName"
                    ]
                    if len(skill_name) > 0:
                        value[i] = skill_name.values[0]
            elif key == "isLoadingBin":
                # 弓箭瓶子处理
                bottle_names = process_loading_bin(value, enum_internal)
                value = bottle_names
            elif key == "Wp08BinType":
                # 斩斧瓶子
                value = get_slash_axe_bin_name(value)
            elif key == "Wp05MusicSkillHighFreqType":
                # 笛子
                v = whistle_high_freq_data.loc[
                    whistle_high_freq_data["HighFreqType"] == value, "SkillName"
                ]
                if len(v) > 0:
                    value = v.values[0]
            elif key == "Wp05HibikiSkillType":
                # 笛子
                v = whistle_hibiki_data.loc[
                    whistle_hibiki_data["HiblkiSkillType"] == value, "SkillName"
                ]
                if len(v) > 0:
                    value = v.values[0]

            if is_guid_like(str(value)):
                # process guids
                text = text_db.get_text_by_guid(value)
                if text:
                    value = text.replace("\n", "").replace("\r", "")
                else:
                    value = ""

            # 处理SlotLevel
            if key == "SlotLevel":
                for i, level in enumerate(value):
                    if level.find("NONE") != -1:
                        value[i] = 0
                    elif level.find("Lv1") != -1:
                        value[i] = 1
                    elif level.find("Lv2") != -1:
                        value[i] = 2
                    elif level.find("Lv3") != -1:
                        value[i] = 3
                    else:
                        print(f"Unknown level: {level}")

            row[key] = value
        table.append(row)

    df = pd.DataFrame(table)
    # 拆分skill列
    skill_names = df["Skill"]
    skill_levels = df["SkillLevel"]
    all_skills = []
    for i in range(len(skill_names)):
        skills = []
        for j in range(len(skill_names[i])):
            level = skill_levels[i][j]
            if level == 0:
                skills.append(None)
            else:
                skills.append(f"{skill_names[i][j]}: {level}")
        skills = list(filter(lambda x: x is not None, skills))
        all_skills.append(skills)
    df["SkillAndLevel"] = all_skills

    df.drop(columns=["Skill", "SkillLevel"], inplace=True)
    df = reindex_column(df, "SkillAndLevel", next_to="SlotLevel")

    df = reindex_column(df, ["ModelId", "CustomModelId"], to_end=True)

    # 处理笛子旋律
    if weapon_type == "whistle" and not keep_serial_id:
        parser = ToneParser()
        parser.set_text_db(text_db)
        parser.load_tone_table(
            "natives/STM/GameDesign/Player/ActionData/Wp05/UserData/Wp05MusicSkillToneTable.user.3.json"
        )
        parser.load_tone_color_table(
            "natives/STM/GameDesign/Player/ActionData/Wp05/UserData/Wp05MusicSkillToneColorTable.user.3.json"
        )
        parser.load_music_skill_data(
            "natives/STM/GameDesign/Common/Player/ActionGuide/MusicSkillData_Wp05.user.3.json"
        )
        parser.set_whistle_data(df)
        df = parser.parse()
        df.drop(
            columns=[
                "MusicSkills",
                "Wp05UniqueType",
            ],
            inplace=True,
        )
        df = reindex_column(df, "MusicSkillNames", next_to="SkillAndLevel")

    # 弩炮弹药处理
    if weapon_type in {"heavybowgun", "lightbowgun"}:
        df["ShellName"] = None
        shell_lv_list_all = df["ShellLv"]
        for row_idx, shell_lv_list in shell_lv_list_all.items():
            shell_name_list = process_gun_shell(
                shell_lv_list, enum_internal, keep_none_level=True
            )
            df.at[row_idx, "ShellName"] = shell_name_list

        # 合并Shell名字，等级和数量列
        def _extract_shell_lv(lv_name: str) -> int:
            if lv_name.startswith("SL_"):
                return int(lv_name[3:]) + 1
            else:
                return 0

        df["MergedShellInfo"] = None
        for row_idx in df.index:
            shell_name_list = df.at[row_idx, "ShellName"]
            shell_lv_list = df.at[row_idx, "ShellLv"]
            shell_num_list = df.at[row_idx, "ShellNum"]
            is_rappid_list = [False for _ in shell_lv_list]
            if "IsRappid" in df.columns:
                # 轻弩速射
                is_rappid_list = df.at[row_idx, "IsRappid"]
            merged = []
            for shell_name, shell_lv, shell_num, is_rappid in zip(
                shell_name_list, shell_lv_list, shell_num_list, is_rappid_list
            ):
                if shell_name is None or shell_num <= 0:
                    continue
                formatted = f"{shell_name} Lv{_extract_shell_lv(shell_lv)} x{shell_num}"
                if is_rappid:
                    formatted += " ↑"
                merged.append(formatted)
            df.at[row_idx, "MergedShellInfo"] = merged

        df = reindex_column(df, "MergedShellInfo", next_to="MainShell")
        df.drop(columns=["ShellName", "ShellLv", "ShellNum"], inplace=True)
        if "IsRappid" in df.columns:
            df.drop(columns=["IsRappid"], inplace=True)

    return df


def dump_weapon_data(keep_serial_id: bool = False) -> dict[str, pd.DataFrame]:
    weapon_types = get_weapon_types()
    weapon_types_lower = {}
    for key, value in weapon_types.items():
        weapon_types_lower[key.lower()] = value

    weapon_paths = {}
    for weapon_type in weapon_types.keys():
        weapon_paths[weapon_type] = (
            f"natives/STM/GameDesign/Common/Weapon/{weapon_type}.user.3.json"
        )

    skill_common_data = dump_skill_common_data(
        "natives/STM/GameDesign/Common/Equip/SkillCommonData.user.3.json"
    )
    enum_internal = load_enum_internal()

    whistle_high_freq_data = dump_user3_data_general(
        "natives/STM/GameDesign/Common/Player/ActionGuide/HighFreqData_Wp05.user.3.json",
        "app.user_data.HighFreqData_Wp05",
    )
    whistle_hibiki_data = dump_user3_data_general(
        "natives/STM/GameDesign/Common/Player/ActionGuide/HibikiData_Wp05.user.3.json",
        "app.user_data.HibikiData_Wp05",
    )

    sheets = {}
    for weapon_type, path in weapon_paths.items():
        df = _dump_weapon_data(
            weapon_paths[weapon_type],
            weapon_type.lower(),
            weapon_types_lower,
            skill_common_data,
            enum_internal,
            whistle_high_freq_data=whistle_high_freq_data,
            whistle_hibiki_data=whistle_hibiki_data,
            keep_serial_id=keep_serial_id,
        )
        sheets[weapon_type] = df
    return sheets


def dump_armor_data() -> pd.DataFrame:
    skill_common_data = dump_skill_common_data(
        "natives/STM/GameDesign/Common/Equip/SkillCommonData.user.3.json"
    )
    armorseries_data = dump_armor_series_data(
        "natives/STM/GameDesign/Common/Equip/ArmorSeriesData.user.3.json"
    )
    armor_data = _dump_armor_data(
        "natives/STM/GameDesign/Common/Equip/ArmorData.user.3.json",
        skill_common_data,
        armorseries_data,
    )
    return armor_data


def _dump_armor_data(
    path: str,
    skill_common_data: pd.DataFrame,
    armor_series_data: pd.DataFrame,
) -> pd.DataFrame:
    data = None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    table = []
    for cData in data[0]["app.user_data.ArmorData"]["_Values"]:
        row = {}
        for key, value in cData["app.user_data.ArmorData.cData"].items():
            if key.startswith("_"):
                key = key[1:]

            value = minify_nested_serial(value)
            value = remove_enum_value(value)
            # 处理Skill
            if key == "Skill":
                for i, skill in enumerate(value):
                    if skill.find("NONE") != -1:
                        continue
                    skill_name = skill_common_data.loc[
                        skill_common_data["skillId"] == skill, "skillName"
                    ]
                    if len(skill_name) > 0:
                        value[i] = skill_name.values[0]

            if is_guid_like(str(value)):
                # process guids
                text = text_db.get_text_by_guid(value)
                if text:
                    value = text.replace("\n", "").replace("\r", "")
                else:
                    value = ""
            # 处理SlotLevel
            if key == "SlotLevel":
                for i, level in enumerate(value):
                    if level == "NONE":
                        value[i] = 0
                    elif level == "Lv1":
                        value[i] = 1
                    elif level == "Lv2":
                        value[i] = 2
                    elif level == "Lv3":
                        value[i] = 3
                    else:
                        print(f"Unknown level: {level}")

            row[key] = value
        table.append(row)

    df = pd.DataFrame(table)
    # Series列改成名字，SeriesId改成原Series
    df["SeriesId"] = df["Series"]
    series_names = []
    for series_id in df["SeriesId"]:
        v = armor_series_data.loc[armor_series_data["Series"] == series_id, "Name"]
        if len(v) > 0:
            series_names.append(v.iloc[0])
        else:
            series_names.append(None)
    df["Series"] = series_names
    # 拆分skill列
    skill_names = df["Skill"]
    skill_levels = df["SkillLevel"]
    all_skills = []
    for i in range(len(skill_names)):
        skills = []
        for j in range(len(skill_names[i])):
            level = skill_levels[i][j]
            if level == 0:
                skills.append(None)
            else:
                skills.append(f"{skill_names[i][j]}: {level}")
        skills = list(filter(lambda x: x is not None, skills))
        all_skills.append(skills)
    df["SkillAndLevel"] = all_skills

    df.drop(columns=["Skill", "SkillLevel"], inplace=True)
    df = reindex_column(df, "SkillAndLevel", next_to="SlotLevel")

    return df


def dump_weapon_series_data(path: str) -> pd.DataFrame:
    data = None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    table = []
    for cData in data[0]["app.user_data.WeaponSeriesData"]["_Values"]:
        cData = cData["app.user_data.WeaponSeriesData.cData"]
        row = {}
        for key, value in cData.items():
            if key.startswith("_"):
                key = key[1:]

            value = minify_nested_serial(value)
            value = remove_enum_value(value)
            text = text_db.get_text_by_guid(str(value))
            if text:
                value = text
            row[key] = value
        table.append(row)
    df = pd.DataFrame(table)
    return df


def dump_armor_series_data(path: str) -> pd.DataFrame:
    data = None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    table = []
    for cData in data[0]["app.user_data.ArmorSeriesData"]["_Values"]:
        cData = cData["app.user_data.ArmorSeriesData.cData"]
        row = {}
        for key, value in cData.items():
            if key.startswith("_"):
                key = key[1:]

            value = minify_nested_serial(value)
            value = remove_enum_value(value)
            text = text_db.get_text_by_guid(str(value))
            if text:
                value = text
            row[key] = value
        table.append(row)
    return pd.DataFrame(table)


if __name__ == "__main__":
    # text_db.set_global_default_lang(1)

    armor_data = dump_armor_data()
    armor_data.drop(columns=["SeriesId"], inplace=True)

    sheets = dump_weapon_data()
    with pd.ExcelWriter("EquipCollection.xlsx") as writer:
        armor_data.to_excel(writer, sheet_name="Armor", index=False)

        for sheet_name, df in sheets.items():
            # 武器表名前加 Wp_ 前缀
            sheet_name = f"Wp_{sheet_name}"
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("Applying autofit...")
        autofit = ExcelAutoFit()
        autofit.style_workbook(writer.book)

        print("Applying rare colors...")
        apply_fix_rare_colors(writer.book)

        align_wrap = Alignment(wrap_text=True)
        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]
            # RARE修正
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    if isinstance(cell.value, str):
                        cell.value = rare_enum_to_value(cell.value)
            # Explain列添加自动换行
            # 查找Explain列
            explain_col = None
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col)
                if cell.value == "Explain":
                    explain_col = col
                    break
            if explain_col is not None:
                for row in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=explain_col)
                    if isinstance(cell.value, str):
                        cell.alignment = align_wrap
