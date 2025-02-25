import json
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from library.excel_auto_fit import ExcelAutoFit
from library.utils import (
    minify_nested_serial,
    reindex_column,
    remove_enum_value,
)
from library.text_db import load_text_db
from table_equip import dump_weapon_series_data, dump_weapon_data, get_weapon_types

text_db = load_text_db("texts_db.json")
re_serial_value = re.compile(r"^\[([-\d]+?)\](.*)$")


# 获取指定GUID的行列位置（绝对位置int，非索引）
def get_pos_of_guid(mat: pd.DataFrame, guid: str) -> tuple[int, int]:
    for i in range(mat.shape[0]):
        for j in range(mat.shape[1]):
            if mat.iat[i, j] == guid:
                return (i, j)
    return None


def sig_set_to_str(sig_set: set) -> str | None:
    if len(sig_set) == 0:
        return None
    if sig_set == {"->"}:
        return "─────"
    if sig_set == {"v"}:
        return "│"
    if sig_set == {"^", "v", "->"}:
        return "   ├───"
    if sig_set == {"^", "->"}:
        return "   └───"

    raise ValueError(f"Invalid sig_set: {sig_set}")


def is_line_cell(text: str) -> bool:
    return text in {"─────", "│", "   ├───", "   └───"}


def dump_weapon_derive_tree(
    path: str,
    weapon_type: str,
    series_data: pd.DataFrame,
    weapon_sheets: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    data = None
    with open(
        path,
        "r",
        encoding="utf-8",
    ) as f:
        data = json.load(f)

    tree_table = []
    series_table = []
    core_data = data[0]["app.user_data.WeaponTree"]
    tree_list = core_data["_WeaponTreeList"]
    for tree_data in tree_list:
        row = {}
        for key, value in tree_data["app.user_data.WeaponTree.cWeaponTree"].items():
            if key.startswith("_"):
                key = key[1:]
            row[key] = value
        tree_table.append(row)

    row_data_list = core_data["_RowDataList"]
    for row_data in row_data_list:
        row = {}
        for key, value in row_data["app.user_data.WeaponTree.cRowData"].items():
            if key.startswith("_"):
                key = key[1:]
            row[key] = value
        series_table.append(row)

    tree_data = pd.DataFrame(tree_table)
    weapon_series_data = pd.DataFrame(series_table)

    weapon_series_data["Series"] = weapon_series_data["Series"].apply(
        lambda x: remove_enum_value(x)
    )

    # {
    #     "_Enable": true,
    #     "_Version": "[0]V_10_00",
    #     "_WeaponID": 1,
    #     "_NextDataGuidList": [
    #         "2347567b-78e0-423d-a64f-919eb14e6502",
    #         "0de6a280-b315-461c-8275-d5cf6fc280b5",
    #         "1fb9d577-5bdf-4b08-ba30-07f673a9ef7e"
    #     ],
    #     "_PreDataGuidList": [],
    #     "_Guid": "e897ff68-c640-4f8d-8cc4-59460e56f048",
    #     "_ColumnDataLevel": 0,
    #     "_RowDataLevel": 0
    # }

    # 建立武器位置矩阵
    col_max = tree_data["ColumnDataLevel"].max()
    row_max = tree_data["RowDataLevel"].max()
    weapon_place_matrix = pd.DataFrame(
        [[None] * (col_max + 1) for _ in range(row_max + 1)],
        columns=range(col_max + 1),
        index=range(row_max + 1),
        dtype="object",
    ).replace(0, None)

    # 填充数据
    for _, row in tree_data.iterrows():
        col_index = row["ColumnDataLevel"]
        col_name = row["RowDataLevel"]
        weapon_place_matrix.at[col_name, col_index] = row["Guid"]

    # 创建位置间隔，用于绘制连线
    for col_name in weapon_place_matrix.columns:
        # 在col_index右方插入 索引为 f"{col_index}_gap" 的空列
        gap_name = f"{col_name}_gap"
        weapon_place_matrix[gap_name] = None
        weapon_place_matrix = reindex_column(
            weapon_place_matrix, column=gap_name, next_to=col_name
        )
    # 移除最后一行和列
    weapon_place_matrix = weapon_place_matrix.iloc[:-1, :-1]
    weapon_place_matrix.replace(np.nan, None, inplace=True)

    # 绘制连线
    col_max = weapon_place_matrix.shape[1]
    row_max = weapon_place_matrix.shape[0]
    line_matrix = pd.DataFrame(
        [[None] * (col_max) for _ in range(row_max)],
        columns=range(col_max),
        index=range(row_max),
        dtype="object",
    )

    def _new_set_or_add(set1: set | None, elem) -> set:
        if set1 is None:
            return {elem}
        set1.add(elem)
        return set1

    for _, data in tree_data.iterrows():
        guid = data["Guid"]
        next_guid_list = data["NextDataGuidList"]
        pos = get_pos_of_guid(weapon_place_matrix, guid)
        for next_guid in next_guid_list:
            next_pos = get_pos_of_guid(weapon_place_matrix, next_guid)
            if not next_pos:
                continue
            # 在两个位置之间创建连线
            down_steps = next_pos[0] - pos[0]
            right_steps = next_pos[1] - pos[1]
            curr_pos = (pos[0], pos[1])
            for _ in range(down_steps):
                # 当前位置修改
                cell = line_matrix.iat[curr_pos[0], curr_pos[1]]
                line_matrix.iat[curr_pos[0], curr_pos[1]] = _new_set_or_add(cell, "v")
                curr_pos = (curr_pos[0] + 1, curr_pos[1])
            for _ in range(right_steps):
                # 当前位置修改
                cell = line_matrix.iat[curr_pos[0], curr_pos[1]]
                line_matrix.iat[curr_pos[0], curr_pos[1]] = _new_set_or_add(cell, "->")
                curr_pos = (curr_pos[0], curr_pos[1] + 1)

    # 优化连线，精确方向
    max_col = line_matrix.shape[1]
    max_row = line_matrix.shape[0]
    for i in range(max_row):
        for j in range(max_col):
            cell = line_matrix.iat[i, j]
            if cell is None:
                continue
            if "->" in cell:
                # 上面的有没有向下箭头
                if i > 0:
                    up_cell = line_matrix.iat[i - 1, j]
                    if up_cell and "v" in up_cell:
                        cell.add("^")
    # 移除重复
    for i in range(line_matrix.shape[0]):
        for j in range(line_matrix.shape[1]):
            if weapon_place_matrix.iat[i, j] is None:
                continue
            line_matrix.iat[i, j] = None

    for i in range(max_row):
        for j in range(max_col):
            cell = line_matrix.iat[i, j]
            if cell is None:
                continue
            line_matrix.iat[i, j] = sig_set_to_str(cell)

    # 应用连线
    for i in range(max_row):
        for j in range(max_col):
            cell = line_matrix.iat[i, j]
            if cell is None:
                continue
            weapon_place_matrix.iat[i, j] = cell

    # 应用名字
    weapon_data = weapon_sheets[weapon_type]

    def _extract_serial_value(text: str) -> str:
        match = re_serial_value.match(text)
        if match:
            return int(match.group(1))
        return -1

    weapon_data["IdEnumValue"] = weapon_data["Id"].apply(_extract_serial_value)

    for i in range(max_row):
        for j in range(max_col):
            prob_guid = weapon_place_matrix.iat[i, j]
            weapon_index = tree_data.loc[tree_data["Guid"] == prob_guid, "WeaponID"]
            if len(weapon_index) == 0:
                continue
            weapon_index = weapon_index.iloc[0]
            name = weapon_data.loc[weapon_data["IdEnumValue"] == weapon_index, "Name"]
            if len(name) == 0:
                continue
            name = name.iloc[0]
            weapon_place_matrix.iat[i, j] = name

    # 应用系列名作为行索引
    for index in weapon_place_matrix.index:
        series_enum = weapon_series_data.loc[
            weapon_series_data["RowLevel"] == int(index), "Series"
        ]
        if len(series_enum) == 0:
            continue
        series_enum = series_enum.iloc[0]
        series_name = series_data.loc[series_data["Series"] == series_enum, "Name"]
        if len(series_name) == 0:
            continue
        series_name = series_name.iloc[0]
        weapon_place_matrix.rename(index={index: series_name}, inplace=True)

    return weapon_place_matrix


if __name__ == "__main__":
    # text_db.set_global_default_lang(1)
    
    weapon_types = get_weapon_types()

    weapon_series_data = dump_weapon_series_data(
        "natives/STM/GameDesign/Common/Equip/WeaponSeriesData.user.3.json"
    )
    weapon_sheets_with_serial_id = dump_weapon_data(keep_serial_id=True)

    with pd.ExcelWriter("WeaponCraftTree.xlsx", engine="openpyxl") as writer:
        for weapon_type in weapon_types.keys():
            print(f"Dumping {weapon_type} tree data...")
            weapon_place_matrix = dump_weapon_derive_tree(
                f"natives/STM/GameDesign/Common/Weapon/{weapon_type}Tree.user.3.json",
                weapon_type,
                weapon_series_data,
                weapon_sheets_with_serial_id,
            )
            weapon_place_matrix.to_excel(writer, sheet_name=weapon_type, index=True)

        # 格式化
        side_style = Side(border_style="thin", color="000000")
        border_style = Border(
            left=side_style, right=side_style, top=side_style, bottom=side_style
        )
        # fill_style = PatternFill(fill_type="solid", fgColor="C5C5C5")
        weapon_sheets = dump_weapon_data(keep_serial_id=False)
        for sheet_name in writer.sheets:
            print(f"Formatting {sheet_name}...")
            weapon_sheet = weapon_sheets[sheet_name]
            sheet = writer.book[sheet_name]
            max_cols = sheet.max_column
            max_rows = sheet.max_row

            # 设置签名
            sheet.cell(row=max_rows + 1, column=1).value = (
                "Created by E - AUTO-GENERATED SHEETS"
            )

            # 设置宽高
            for i in range(max_cols):
                if i == 0:
                    sheet.column_dimensions[get_column_letter(i + 1)].width = 15
                    continue
                sheet.column_dimensions[get_column_letter(i + 1)].width = 10
            for i in range(max_rows):
                sheet.row_dimensions[i + 1].height = 30
            # 去掉列名行的临时数据
            for i in range(max_cols):
                cell = sheet.cell(row=1, column=i + 1)
                if isinstance(cell.value, str) and cell.value.find("_gap") != -1:
                    cell.value = ""
            # 设置字体和居中
            for i in range(2, max_rows + 1):
                for j in range(2, max_cols + 1):
                    cell = sheet.cell(row=i, column=j)
                    if isinstance(cell.value, str) and is_line_cell(cell.value):
                        # 连线单元格
                        cell.font = Font(name="Courier New", size=11)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    else:
                        # 正文所有单元格
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )
                        if isinstance(cell.value, str) and cell.value.strip() != "":
                            # 非空单元格
                            cell.border = border_style
                            # cell.fill = fill_style
                            # 判断武器属性
                            attr = weapon_sheet.loc[
                                weapon_sheet["Name"] == cell.value, "Attribute"
                            ]
                            if len(attr) > 0:
                                attr = attr.iloc[0]
                                if attr == "FIRE":
                                    cell.font = Font(color="E15057")
                                    cell.value = f"{cell.value} (火)"
                                elif attr == "WATER":
                                    cell.font = Font(color="6CA3D9")
                                    cell.value = f"{cell.value} (水)"
                                elif attr == "ICE":
                                    cell.font = Font(color="2EC9E6")
                                    cell.value = f"{cell.value} (冰)"
                                elif attr == "ELEC":
                                    cell.font = Font(color="F2C21D")
                                    cell.value = f"{cell.value} (雷)"
                                elif attr == "DRAGON":
                                    cell.font = Font(color="56379E")
                                    cell.value = f"{cell.value} (龙)"
                                elif attr == "PARALYSE":
                                    cell.font = Font(color="B38F48")
                                    cell.value = f"{cell.value} (麻)"
                                elif attr == "SLEEP":
                                    cell.font = Font(color="685ECD")
                                    cell.value = f"{cell.value} (眠)"
                                elif attr == "POISON":
                                    cell.font = Font(color="9788D1")
                                    cell.value = f"{cell.value} (毒)"
