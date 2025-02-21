import openpyxl
import re


class ExcelAutoFit:
    RE_HAN = re.compile(r"[\u4e00-\u9fa5]")
    MIN_WIDTH = 6.0
    MAX_WIDTH = 80.0

    # 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
    def get_num_colnum_dict(self):
        """
        :return: 返回字典：{1:'A', 2:'B', ...... , 52:'AZ'}
        """
        num_str_dict = {}
        A_Z = [chr(a) for a in range(ord("A"), ord("Z") + 1)]
        AA_AZ = ["A" + chr(a) for a in range(ord("A"), ord("Z") + 1)]
        A_AZ = A_Z + AA_AZ
        for i in A_AZ:
            num_str_dict[A_AZ.index(i) + 1] = i
        return num_str_dict

    # 自适应列宽
    def style_excel(self, excel_name: str):
        """
        :param sheet_name:  excel中的sheet名
        :return:
        """
        # 打开excel
        wb = openpyxl.load_workbook(excel_name)
        # 选择对应的sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # 获取最大行数与最大列数
            max_column = sheet.max_column
            max_row = sheet.max_row

            # 将每一列，单元格列宽最大的列宽值存到字典里，key:列的序号从1开始(与字典num_str_dic中的key对应)；value:列宽的值
            max_column_dict = {}

            # 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
            num_str_dict = self.get_num_colnum_dict()

            # 遍历全部列
            for i in range(1, max_column + 1):
                # 遍历每一列的全部行
                for j in range(1, max_row + 1):
                    column = 0
                    # 获取j行i列的值
                    sheet_value = sheet.cell(row=j, column=i).value
                    # 通过列表生成式生成字符列表，将当前获取到的单元格的str值的每一个字符放在一个列表中（列表中一个元素是一个字符）
                    sheet_value_list = [k for k in str(sheet_value)]
                    # 遍历当前单元格的字符列表
                    for v in sheet_value_list:
                        # 判定长度，一个数字或一个字母，单元格列宽+=1.1，其它+=2.2（长度可根据需要自行修改，经测试一个字母的列宽长度大概为1）
                        vb = v.encode("utf-8")
                        if vb.isdigit() or vb.isalpha():
                            column += 1.2
                        elif self.RE_HAN.search(v):
                            column += 2.0
                        else:
                            column += 1.0
                    # 当前单元格列宽与字典中的对比，大于字典中的列宽值则将字典更新。如果字典没有这个key，抛出异常并将值添加到字典中
                    try:
                        if column > max_column_dict[i]:
                            max_column_dict[i] = column
                    except Exception as e:
                        max_column_dict[i] = column
            # 此时max_column_dict字典中已存有当前sheet的所有列的最大列宽值，直接遍历字典修改列宽
            for key, value in max_column_dict.items():
                if value < self.MIN_WIDTH:
                    value = self.MIN_WIDTH
                elif value > self.MAX_WIDTH:
                    value = self.MAX_WIDTH
                try:
                    sheet.column_dimensions[num_str_dict[key]].width = value
                except Exception as e:
                    print(f"AutoFit error: {e}, key: {key}, value: {value}")
        # 保存
        wb.save(excel_name)


if __name__ == "__main__":
    # 调用方法 实例化类
    Entity = ExcelAutoFit()
    # 传入参数：Excel名称，需要设置列宽的Sheet名称
    Entity.style_excel("MissionDataCollection.xlsx")
