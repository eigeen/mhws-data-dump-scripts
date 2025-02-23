import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color

RARE_COLORS = {
    0: "969696",
    1: "DEDEDE",
    2: "A4C43B",
    3: "47A33F",
    4: "5CAEBB",
    5: "575FD9",
    6: "9272E3",
    7: "C76D46",
    8: "B3436A",
    9: "2EC9E6",
    10: "F2C21D",
    11: "B4F5FF",
}

re_rare = re.compile(r"^RARE(\d+)$")


# 计算颜色的亮度
def _get_brightness(color):
    if isinstance(color, Color):
        color = color.rgb
    r = int(color[0:2], 16)
    g = int(color[2:4], 16)
    b = int(color[4:6], 16)
    return (r * 299 + g * 587 + b * 114) / 1000


# 根据亮度选择文字颜色
def _get_text_color(brightness):
    return Font(color="FFFFFF") if brightness < 64 else Font(color="000000")


# 混合颜色以模拟不透明度
def _apply_opacity(color, opacity):
    """
    将颜色与白色混合以模拟不透明度。
    :param color: 十六进制颜色代码（如 "FF0000"）
    :param opacity: 不透明度（0.0 到 1.0，0.0 为完全透明，1.0 为完全不透明）
    :return: 混合后的十六进制颜色代码
    """
    if opacity < 0 or opacity > 1:
        raise ValueError("Opacity must be between 0 and 1")

    # 将颜色和白色转换为 RGB
    r = int(color[0:2], 16)
    g = int(color[2:4], 16)
    b = int(color[4:6], 16)
    white_r, white_g, white_b = 255, 255, 255

    # 混合颜色
    mixed_r = int(r * opacity + white_r * (1 - opacity))
    mixed_g = int(g * opacity + white_g * (1 - opacity))
    mixed_b = int(b * opacity + white_b * (1 - opacity))

    # 返回混合后的十六进制颜色
    return f"{mixed_r:02X}{mixed_g:02X}{mixed_b:02X}"


def apply_rare_colors(wb: Workbook):
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        max_column = ws.max_column
        max_row = ws.max_row
        for i in range(max_row):
            for j in range(max_column):
                cell = ws.cell(row=i + 1, column=j + 1)
                if cell.value is not None and isinstance(cell.value, str):
                    match = re_rare.search(cell.value)
                    if match:
                        color = RARE_COLORS[int(match.group(1))]
                        color = _apply_opacity(color, 0.5)
                        brightness = _get_brightness(Color(color))
                        text_color = _get_text_color(brightness)
                        cell.font = text_color
                        cell.fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )
        # cell = ws.cell(row=i+1, column=1, value=f"Background: {color}")
        # fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # cell.fill = fill
