import json
import math
import re
import pandas as pd

import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

from library.excel_auto_fit import ExcelAutoFit
from library.item_db import ItemDB
from library.text_db import load_text_db
from library.utils import is_guid_like, remove_enum_value, reindex_column

item_db = ItemDB("item_db.json")
text_db = load_text_db("texts_db.json")

re_rare = re.compile(r"^RARE(\d+)$")


def minify_nested_obj(obj: dict) -> str | dict:
    if isinstance(obj, dict) and len(obj) == 1:
        first_key = list(obj.keys())[0]
        value = obj[first_key]
        if "_Value" in value:
            return value["_Value"]
        for k, v in value.items():
            value[k] = minify_nested_obj(v)

    if isinstance(obj, list):
        for i in range(len(obj)):
            obj[i] = minify_nested_obj(obj[i])

    return obj


def dump_skill_common_data(path: str) -> pd.DataFrame:
    data = None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    table = []
    for cData in data[0]["app.user_data.SkillCommonData"]["_Values"]:
        cData = cData["app.user_data.SkillCommonData.cData"]

        row = {}
        for key, value in cData.items():
            if key.startswith("_"):
                key = key[1:]
            value = minify_nested_obj(value)
            value = remove_enum_value(value)
            if is_guid_like(str(value)):
                text = text_db.get_text_by_guid(value)
                if text:
                    value = text.replace("\n", "").replace("\r", "")
                    if value.startswith("<COLOR FF0000>#Rejected#</COLOR> "):
                        value = value[33:]
            row[key] = value
        table.append(row)

    df = pd.DataFrame(table)
    return df


def dump_skill_data(path: str, skill_common_data: pd.DataFrame) -> pd.DataFrame:
    data = None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    table = []
    for cData in data[0]["app.user_data.SkillData"]["_Values"]:
        cData = cData["app.user_data.SkillData.cData"]

        row = {}
        for key, value in cData.items():
            if key.startswith("_"):
                key = key[1:]
            value = minify_nested_obj(value)
            value = remove_enum_value(value)

            if key == "openSkill":
                for i, skill_id in enumerate(value):
                    if skill_id != "NONE":
                        try:
                            skill_name = skill_common_data[
                                skill_common_data["skillId"] == skill_id
                            ]["skillName"].iloc[0]
                            value[i] = skill_name
                        except:
                            pass
            if is_guid_like(str(value)):
                text = text_db.get_text_by_guid(value)
                if text:
                    value = text.replace("\n", "").replace("\r", "")
                    if value.startswith("<COLOR FF0000>#Rejected#</COLOR> "):
                        value = value[33:]
            row[key] = value
        table.append(row)

    df = pd.DataFrame(table)
    return df


def dump_accessory_data(path: str, skill_common_data: pd.DataFrame) -> pd.DataFrame:
    data = None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    table = []
    for cData in data[0]["app.user_data.AccessoryData"]["_Values"]:
        cData = cData["app.user_data.AccessoryData.cData"]

        row = {}
        for key, value in cData.items():
            if key.startswith("_"):
                key = key[1:]
            value = minify_nested_obj(value)
            value = remove_enum_value(value)

            if key == "Skill":
                for i, skill_id in enumerate(value):
                    if skill_id != "NONE":
                        skill_name = skill_common_data[
                            skill_common_data["skillId"] == skill_id
                        ]["skillName"].iloc[0]
                        value[i] = skill_name
            if key == "Rare":
                match = re_rare.match(value)
                if match:
                    value = int(match.group(1)) + 1 
            if is_guid_like(str(value)):
                text = text_db.get_text_by_guid(value)
                if text:
                    value = text.replace("\n", "").replace("\r", "")
            else:
                item = item_db.get_entry_by_id(str(value))
                if item:
                    value = item.raw_name
            row[key] = value
        table.append(row)

    df = pd.DataFrame(table)
    return df


def dump_accessory_ratio_data(
    judge_path: str, rank_judge_path: str, accessory_data: pd.DataFrame
) -> pd.DataFrame:
    judge_data = None
    with open(judge_path, "r", encoding="utf-8") as f:
        judge_data = json.load(f)

    rank_judge_data = None
    with open(rank_judge_path, "r", encoding="utf-8") as f:
        rank_judge_data = json.load(f)

    acc_prob_table = []
    for cData in judge_data[0]["app.user_data.AccessoryJudgeData"]["_Values"]:
        cData = cData["app.user_data.AccessoryJudgeData.cData"]
        row = {}
        for key, value in cData.items():
            if key.startswith("_"):
                key = key[1:]
            value = minify_nested_obj(value)
            value = remove_enum_value(value)
            row[key] = value
        acc_prob_table.append(row)

    rank_prob_table = []
    for cData in rank_judge_data[0]["app.user_data.AccessoryRankJudgeData"]["_Values"]:
        cData = cData["app.user_data.AccessoryRankJudgeData.cData"]
        row = {}
        for key, value in cData.items():
            if key.startswith("_"):
                key = key[1:]
            value = minify_nested_obj(value)
            value = remove_enum_value(value)

            # item = item_db.get_entry_by_id(str(value))
            # if item:
            #     value = item.raw_name
            row[key] = value
        rank_prob_table.append(row)

    sword_pool_weight_sums = [0, 0, 0, 0, 0]  # 每个池的权重之和
    equip_pool_weight_sums = [0, 0, 0, 0, 0]
    for acc_row in acc_prob_table:
        # 获取珠子类别
        acc_type = accessory_data.loc[
            accessory_data["AccessoryId"] == acc_row["AccessoryId"], "AccessoryType"
        ].iloc[0]
        for i in range(len(acc_row["Probability"])):
            if acc_type.find("ACC_TYPE_00") != -1:
                sword_pool_weight_sums[i] += acc_row["Probability"][i]
            elif acc_type.find("ACC_TYPE_01") != -1:
                equip_pool_weight_sums[i] += acc_row["Probability"][i]
            else:
                print("unknown accessory type: ", acc_type)
    print("sword_pool_weight_sums: ", sword_pool_weight_sums)
    print("equip_pool_weight_sums: ", equip_pool_weight_sums)

    out_data = {}
    for acc_row in acc_prob_table:
        row = {}
        # 获取珠子类别
        acc_type = accessory_data.loc[
            accessory_data["AccessoryId"] == acc_row["AccessoryId"], "AccessoryType"
        ].iloc[0]
        pool_weight_sums = None
        if acc_type.find("ACC_TYPE_00") != -1:
            pool_weight_sums = sword_pool_weight_sums
        elif acc_type.find("ACC_TYPE_01") != -1:
            pool_weight_sums = equip_pool_weight_sums
        else:
            print("unknown accessory type: ", acc_type)
            continue
        for rank_row in rank_prob_table:
            if rank_row["AccessoryType"] != acc_type:
                row[rank_row["ItemId"]] = 0
                continue
            acc_ratio = 0
            for pool_idx, pool_weight in enumerate(acc_row["Probability"]):
                if pool_weight_sums[pool_idx] == 0:
                    continue
                pool_ratio = pool_weight / pool_weight_sums[pool_idx]
                rank_ratio = rank_row["Probability"][pool_idx] / sum(
                    rank_row["Probability"]
                )
                acc_ratio += pool_ratio * rank_ratio
            row[rank_row["ItemId"]] = acc_ratio
        out_data[acc_row["AccessoryId"]] = row

    df = pd.DataFrame(out_data).T
    return df


def export_percent_pretty(accessory_percent_data: pd.DataFrame):
    # 导出美化的珠子爆率表
    accessory_prob_pretty_data = accessory_percent_data.copy()
    accessory_prob_pretty_data.drop(
        ["Index", "AccessoryType", "SortId", "Price", "Skill", "SkillLevel"],
        axis=1,
        inplace=True,
    )
    for col_name in accessory_prob_pretty_data.columns:
        if col_name.startswith("Prob: "):
            accessory_prob_pretty_data.rename(
                columns={col_name: col_name[6:]}, inplace=True
            )
    accessory_prob_pretty_data = reindex_column(
        accessory_prob_pretty_data,
        column=["技能1", "技能1等级", "技能2", "技能2等级"],
        next_to="Rare",
    )
    # 技能为NONE和等级为0的设为None值
    accessory_prob_pretty_data[["技能1", "技能1等级", "技能2", "技能2等级"]] = (
        accessory_prob_pretty_data[
            ["技能1", "技能1等级", "技能2", "技能2等级"]
        ].replace(to_replace=["NONE", 0], value=None)
    )
    # 排序 期望值↓ 稀有度↑ Id↑
    accessory_prob_pretty_data.sort_values(
        by=["期望值", "Rare", "AccessoryId"],
        ascending=[False, True, True],
        inplace=True,
    )
    accessory_prob_pretty_data.rename(
        columns={
            "技能1等级": "1等级",
            "技能2等级": "2等级",
            "AccessoryId": "Id",
            "Rare": "稀有度",
            "期望值": "期望（N出1）",
        },
        inplace=True,
    )
    # print(accessory_prob_pretty_data)

    autofit = ExcelAutoFit()
    with pd.ExcelWriter("AccessoryPercentPretty.xlsx", engine="openpyxl") as writer:
        accessory_prob_pretty_data.to_excel(
            writer, sheet_name="AccessoryPercent", index=False
        )
        # 修订格式
        percent3_style = NamedStyle(name="percent3_style", number_format="0.000%")
        sheet = writer.book.active
        # 遍历列头
        for col in sheet.iter_cols(1, sheet.max_column):
            column_header = col[0].value  # 获取列头的值
            if column_header in {
                "谜之宝珠·剑",
                "发光宝珠·剑",
                "老旧宝珠·剑",
                "谜之宝珠·铠",
                "发光宝珠·铠",
                "老旧宝珠·铠",
            }:
                col_letter = get_column_letter(col[0].column)  # 获取列字母
                for cell in sheet[col_letter]:
                    if isinstance(cell.value, (int, float)):  # 检查单元格的值是否为数字
                        cell.style = percent3_style  # 应用样式

        autofit.style_workbook(writer.book)


if __name__ == "__main__":
    # text_db.set_global_default_lang(13)

    skill_common_data = dump_skill_common_data(
        "natives/STM/GameDesign/Common/Equip/SkillCommonData.user.3.json"
    )
    accessory_data = dump_accessory_data(
        "natives/STM/GameDesign/Common/Equip/AccessoryData.user.3.json",
        skill_common_data,
    )
    accessory_ratio_data = dump_accessory_ratio_data(
        "natives/STM/GameDesign/Common/Equip/AccessoryJudgeData.user.3.json",
        "natives/STM/GameDesign/Common/Equip/AccessoryRankJudgeData.user.3.json",
        accessory_data,
    )
    # 根据ACC_ID合并珠子抽卡概率
    accessory_data = accessory_data.merge(
        accessory_ratio_data, left_on="AccessoryId", right_index=True
    )

    # 替换索引的Acc ID为名称
    for col_name in accessory_data.columns:
        name = item_db.get_entry_by_id(col_name)
        if name:
            accessory_data.rename(
                columns={col_name: f"Prob: {name.raw_name}"}, inplace=True
            )
    # print(accessory_data)

    sheets = {
        "SkillCommonData": skill_common_data,
        "SkillData": dump_skill_data(
            "natives/STM/GameDesign/Common/Equip/SkillData.user.3.json",
            skill_common_data,
        ),
        "AccessoryData": accessory_data,
    }

    autofit = ExcelAutoFit()
    with pd.ExcelWriter("SkillCollection.xlsx", engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        autofit.style_workbook(writer.book)

        # 修订格式
        percent3_style = NamedStyle(name="percent3_style", number_format="0.000%")
        sheet = writer.book["AccessoryData"]
        # 遍历列头
        for col in sheet.iter_cols(1, sheet.max_column):
            column_header = col[0].value  # 获取列头的值
            if column_header.startswith("Prob: "):
                col_letter = get_column_letter(col[0].column)  # 获取列字母
                for cell in sheet[col_letter]:
                    if isinstance(cell.value, (int, float)):  # 检查单元格的值是否为数字
                        cell.style = percent3_style  # 应用样式

    # 导出额外分表
    accessory_percent_data = accessory_data.copy()
    accessory_percent_data.drop(
        ["Explain", "IconColor", "SlotLevelAcc"], axis=1, inplace=True
    )
    skills = accessory_percent_data["Skill"]
    skill_levels = accessory_percent_data["SkillLevel"]
    skill_data = []
    for i in range(len(skills)):
        acc_skills = []
        for j in range(len(skills[i])):
            acc_skills.append([skills[i][j], skill_levels[i][j]])
        skill_data.append(acc_skills)
    accessory_percent_data["技能1"] = list(map(lambda x: x[0][0], skill_data))
    accessory_percent_data["技能1等级"] = list(map(lambda x: x[0][1], skill_data))
    accessory_percent_data["技能2"] = list(map(lambda x: x[1][0], skill_data))
    accessory_percent_data["技能2等级"] = list(map(lambda x: x[1][1], skill_data))
    accessory_percent_data["期望值"] = (
        accessory_percent_data.filter(like="Prob: ")
        # .sum(axis=1)
        .max(axis=1)
        .apply(lambda x: 1 / x if x != 0 else 0)
        .apply(lambda x: math.ceil(x))
    )
    # 遍历所有单元格，移除enum value prefix [123]ENUM_VALUE -> ENUM_VALUE
    for col in accessory_percent_data.columns:
        for idx, cell in enumerate(accessory_percent_data[col]):
            cell_new = remove_enum_value(cell)
            if cell_new != cell:
                accessory_percent_data.at[idx, col] = cell_new

    with pd.ExcelWriter("AccessoryPercent.xlsx", engine="openpyxl") as writer:
        accessory_percent_data.to_excel(
            writer, sheet_name="AccessoryPercent", index=False
        )
    autofit.style_excel("AccessoryPercent.xlsx")

    # accessory_ratio_data_raw = accessory_ratio_data.copy()
    # for col in accessory_ratio_data_raw.columns:
    #     col_new = remove_enum_value(col)
    #     if col_new != col:
    #         accessory_ratio_data_raw.rename(columns={col: col_new}, inplace=True)
    # for idx in accessory_ratio_data_raw.index:
    #     col_new = remove_enum_value(col)
    #     if col_new != col:
    #         accessory_ratio_data_raw.rename(columns={col: col_new}, inplace=True)
    # accessory_ratio_data_raw.to_csv("AccessoryPercentRaw.csv")

    export_percent_pretty(accessory_percent_data)
