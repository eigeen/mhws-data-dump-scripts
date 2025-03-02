import pandas as pd
import os

import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from library.excel_auto_fit import ExcelAutoFit
from library.utils import rare_enum_to_value, reindex_column, seperate_enum_value
from library.image_utils import compress_png
from table_equip import (
    dump_armor_data,
    dump_weapon_data,
    get_weapon_types,
    dump_armor_series_enum_maker,
)
from table_general import create_icon_flag, parse_icon_flag


def apply_icons(
    workbook: openpyxl.Workbook,
    icon_dir: str,
    skip_row: int = 0,
):
    for sheet_name in workbook.sheetnames:
        sheet = writer.book[sheet_name]

        # 遍历所有单元格，查找icon flag
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2)):
            for col_idx, cell in enumerate(row):
                if not isinstance(cell.value, str):
                    continue
                icon_flag = parse_icon_flag(cell.value)
                if not icon_flag:
                    continue
                icon_file_name = icon_flag["path"]
                if not icon_file_name:
                    continue
                # 调整列宽
                sheet.column_dimensions[get_column_letter(col_idx + 1)].width = 13
                # 读取文件并插入图片
                icon_path = os.path.join(
                    icon_dir,
                    icon_file_name,
                )
                # 压缩图片
                try:
                    icon_path = compress_png(icon_path)
                except Exception as e:
                    print(f"Error compressing {icon_path}: {e}")
                    continue

                if os.path.exists(icon_path):
                    img = OpenpyxlImage(icon_path)
                    img.width = 100
                    img.height = 100
                    cell.value = ""
                    # 将图片插入到当前格
                    sheet.add_image(
                        img,
                        get_column_letter(col_idx + 1) + str(cell.row),
                    )
                    # 调整行高
                    sheet.row_dimensions[cell.row].height = 80
                else:
                    print(f"File not found: {icon_path}")


def append_icon_col_weapon(
    weapon_data: pd.DataFrame,
    weapon_data_serial: pd.DataFrame,
    weapon_id_int: int,
) -> pd.DataFrame:
    # 提取Id enum value
    ids = []
    for _, row in weapon_data_serial.iterrows():
        int_id, _ = seperate_enum_value(row["Id"])
        ids.append(int_id)
    icon_flags = []
    for icon_id in ids:
        icon_file_name = (
            f"tex_it{weapon_id_int:02d}00_0{icon_id:03d}_IMLM4.tex.241106027.png"
        )
        icon_flag = create_icon_flag(icon_file_name)
        icon_flags.append(icon_flag)
    # 添加Icon列
    weapon_data["Icon"] = icon_flags
    weapon_data = reindex_column(weapon_data, "Icon", next_to="Name")

    return weapon_data


def append_icon_col_armor(
    armor_data: pd.DataFrame, armor_series_enum_maker: pd.DataFrame
) -> pd.DataFrame:
    # 提取series id
    series_ids = []
    for _, row in armor_data.iterrows():
        series_id_enum = row["SeriesId"]
        v = armor_series_enum_maker.loc[
            armor_series_enum_maker["EnumName"] == series_id_enum, "EnumValue"
        ]
        if len(v) == 0:
            print(f"Invalid series id: {series_id_enum}")
            series_ids.append(0)
        else:
            series_ids.append(int(v.values[0]))
    parts_types = armor_data["PartsType"].tolist()
    parts_type_dict = {
        "HELM": 0,
        "BODY": 1,
        "ARM": 2,
        "WAIST": 3,
        "LEG": 4,
    }
    icon_flags = {0: [], 1: []}
    for series_id, parts_type in zip(series_ids, parts_types):
        for ab_group in [0, 1]:
            if series_id in {0, 1}:
                icon_flags[ab_group].append("")
                continue
            icon_file_name = f"tex_ch03_00_{series_id:03d}_{ab_group:01d}_{parts_type_dict[parts_type]}_IMLM4.tex.241106027.png"
            icon_flag = create_icon_flag(icon_file_name)
            icon_flags[ab_group].append(icon_flag)
    # 添加Icon列
    armor_data["Icon_A"] = icon_flags[0]
    armor_data["Icon_B"] = icon_flags[1]
    armor_data = reindex_column(armor_data, ["Icon_A", "Icon_B"], next_to="Name")
    armor_data.drop(columns=["SeriesId"], inplace=True)
    return armor_data


if __name__ == "__main__":
    print("Dumping weapon data...")
    weapon_sheets_serial = dump_weapon_data(keep_serial_id=True)
    weapon_sheets = dump_weapon_data(keep_serial_id=False)
    weapon_types = get_weapon_types()
    armor_data = dump_armor_data()
    armor_series_enum_maker = dump_armor_series_enum_maker()

    icon_dir = "C:/Users/Eigeen/Downloads/tex"

    autofit = ExcelAutoFit()
    with pd.ExcelWriter("WeaponDataWithIcon.xlsx") as writer:
        for weapon_type, weapon_data in weapon_sheets.items():
            weapon_data = append_icon_col_weapon(
                weapon_data,
                weapon_sheets_serial[weapon_type],
                weapon_types[weapon_type],
            )

            weapon_data.to_excel(writer, sheet_name=weapon_type, index=False)

        print("Styling...")
        autofit.style_workbook(writer.book, max_width=60)

        print("Applying wrap text alignment...")
        align_wrap = Alignment(wrapText=True)
        for sheet_name in writer.sheets:
            sheet = writer.book[sheet_name]
            # RARE修正
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    if isinstance(cell.value, str):
                        cell.value = rare_enum_to_value(cell.value)
            # 为所有非表头格应用自动换行
            for row in sheet.iter_rows(min_row=3):
                for cell in row:
                    cell.alignment = align_wrap
            # # 首行增加一行，加签名
            # sheet.insert_rows(1)
            # sheet.cell(row=1, column=1).value = (
            #     "解包：剧透猎人老金群  制表：Eigeen本征  版本：1.0  不包含首日补丁"
            # )

        print("Applying icons...")
        # apply_icons(writer.book, icon_dir, skip_row=1)
        apply_icons(writer.book, icon_dir)

    with pd.ExcelWriter("ArmorDataWithIcon.xlsx") as writer:
        armor_data = append_icon_col_armor(armor_data, armor_series_enum_maker)
        armor_data.to_excel(writer, sheet_name="Armor", index=False)
        print("Styling...")
        autofit.style_workbook(writer.book, max_width=60)

        print("Applying wrap text alignment...")
        align_wrap = Alignment(wrapText=True)
        for sheet_name in writer.sheets:
            sheet = writer.book[sheet_name]
            # 为所有非表头格应用自动换行
            for row in sheet.iter_rows(min_row=3):
                for cell in row:
                    cell.alignment = align_wrap
            # # 首行增加一行，加签名
            # sheet.insert_rows(1)
            # sheet.cell(row=1, column=1).value = (
            #     "解包：剧透猎人老金群  制表：Eigeen本征  版本：1.0  不包含首日补丁"
            # )

        print("Applying icons...")
        # apply_icons(writer.book, icon_dir, skip_row=1)
        apply_icons(writer.book, icon_dir)
