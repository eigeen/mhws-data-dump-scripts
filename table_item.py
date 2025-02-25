import json
import pandas as pd
from openpyxl.styles import Alignment

from library.excel_auto_fit import ExcelAutoFit
from library.rare import apply_rare_colors
from library.text_db import load_text_db
from library.item_db import ItemDB
from library.utils import is_guid_like, rare_enum_to_value, remove_enum_value
from table_skill import minify_nested_obj

text_db = load_text_db("texts_db.json")


def dump_item_data(path: str) -> pd.DataFrame:
    data = None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    table = []
    for cData in data[0]["app.user_data.ItemData"]["_Values"]:
        cData = cData["app.user_data.ItemData.cData"]

        row = {}
        for key, value in cData.items():
            if key.startswith("_"):
                key = key[1:]
            value = minify_nested_obj(value)
            value = remove_enum_value(value)

            if isinstance(value, str):
                if is_guid_like(value):
                    text = text_db.get_text_by_guid(value)
                    if text:
                        value = text.replace("\n", "").replace("\r", "")
                        if value.startswith("<COLOR FF0000>#Rejected#</COLOR> "):
                            value = value[33:]
                elif value == "INVALID":
                    value = ""
            row[key] = value
        table.append(row)

    df = pd.DataFrame(table)
    return df


def dump_item_recipe_data(path: str, item_data: pd.DataFrame) -> pd.DataFrame:
    data = None
    with open(path, "r") as f:
        data = json.load(f)

    table = []
    for cData in data[0]["app.user_data.cItemRecipe"]["_Values"]:
        cData = cData["app.user_data.cItemRecipe.cData"]

        row = {}
        for key, value in cData.items():
            if key.startswith("_"):
                key = key[1:]

            value = minify_nested_obj(value)
            value = remove_enum_value(value)
            # replace item_id with item_name
            if isinstance(value, list):
                for i, item_id in enumerate(value):
                    v = item_data.loc[item_data["ItemId"] == item_id, "RawName"]
                    if len(v) > 0:
                        value[i] = v.values[0]
                value = ", ".join(value)
            elif isinstance(value, str):
                v = item_data.loc[item_data["ItemId"] == value, "RawName"]
                if len(v) > 0:
                    value = v.values[0]

            row[key] = value

        table.append(row)
    df = pd.DataFrame(table)
    return df


if __name__ == "__main__":
    text_db.set_global_default_lang(1)
    
    item_data = dump_item_data(
        "natives/STM/GameDesign/Common/Item/itemData.user.3.json"
    )
    item_recipe_data = dump_item_recipe_data(
        "natives/STM/GameDesign/Common/Item/ItemRecipe.user.3.json", item_data
    )

    sheets = {
        "ItemData": item_data,
        "ItemRecipeData": item_recipe_data,
    }

    auto_fit = ExcelAutoFit()
    align_wrap = Alignment(wrapText=True)
    with pd.ExcelWriter("ItemDataCollection.xlsx") as writer:
        for sheet_name, data in sheets.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)

        auto_fit.style_workbook(writer.book)
        apply_rare_colors(writer.book)

        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]
            # RARE修正
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    if isinstance(cell.value, str):
                        cell.value = rare_enum_to_value(cell.value)
        # RawExplain列添加自动换行
        explain_col = None
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            if cell.value == "RawExplain":
                explain_col = col
                break
        if explain_col is not None:
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=explain_col)
                if isinstance(cell.value, str):
                    cell.alignment = align_wrap
