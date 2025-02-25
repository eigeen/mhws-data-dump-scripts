import json
import os
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

from library.excel_auto_fit import ExcelAutoFit
from library.text_db import load_text_db
from library.utils import is_guid_like, remove_enum_value, reindex_column

text_db = load_text_db("texts_db.json")


def dump_enemy_data(enemy_path: str, species_path: str) -> pd.DataFrame:
    species_data = dump_species_data(species_path)
    return _dump_enemy_data(enemy_path, species_data)


def _dump_enemy_data(path: str, species_data: pd.DataFrame) -> pd.DataFrame:
    data = None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    table = []
    for cData in data[0]["app.user_data.EnemyData"]["_Values"]:
        row = {}
        for key, value in cData["app.user_data.EnemyData.cData"].items():
            if key.startswith("_"):
                key = key[1:]

            if is_guid_like(str(value)):
                # process guids
                text = text_db.get_text_by_guid(value)
                if text:
                    value = text.replace("\n", "").replace("\r", "")
                else:
                    value = ""
            if key == "Species":
                specie_name = species_data.loc[
                    species_data["EmSpecies"] == value, "EmSpeciesName"
                ]
                if len(specie_name) > 0:
                    value = specie_name.values[0]

            value = remove_enum_value(value)

            row[key] = value
        table.append(row)
    df = pd.DataFrame(table)
    # wtf JpEnemyName
    df = reindex_column(df, "JpEnemyName", to_end=True)

    return df


def dump_species_data(path: str) -> pd.DataFrame:
    data = None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    table = []
    for cData in data[0]["app.user_data.EnemySpeciesData"]["_Values"]:
        row = {}
        for key, value in cData["app.user_data.EnemySpeciesData.cData"].items():
            if key.startswith("_"):
                key = key[1:]
            if is_guid_like(str(value)):
                # process guids
                text = text_db.get_text_by_guid(value)
                if text:
                    value = text.replace("\n", "").replace("\r", "")
                else:
                    value = ""

            row[key] = value
        table.append(row)
    df = pd.DataFrame(table)
    return df


def main():
    # text_db.set_global_default_lang(1)

    species_data = dump_species_data(
        "natives/STM/GameDesign/Common/Enemy/EnemySpecies.user.3.json"
    )
    enemy_data = _dump_enemy_data(
        "natives/STM/GameDesign/Common/Enemy/EnemyData.user.3.json", species_data
    )
    # 统计种族数量
    species_counts = enemy_data["Species"].value_counts()
    species_counts = species_counts.reset_index()
    species_counts.columns = ["EmSpeciesName", "Count"]
    species_data = species_data.merge(species_counts, on="EmSpeciesName")
    # 为图片预留文件名
    enemy_data["Icon"] = enemy_data["enemyId"].apply(
        lambda x: "" if x == "INVALID" else x
    )
    enemy_data = reindex_column(enemy_data, "Icon", next_to="EnemyName")

    sheets = {
        "EnemyData": enemy_data,
        "SpeciesData": species_data,
    }

    autofit = ExcelAutoFit()
    with pd.ExcelWriter("EnemyCollection.xlsx", engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 添加图片 credit: soulize
        ws = writer.book["EnemyData"]

        icon_dir = "em_icons"
        # 遍历第一行（表头），查找Icon列的位置
        icon_col = -1
        for i, col in enumerate(ws.iter_cols()):
            if col[0].value == "Icon":
                icon_col = i
                break
        if icon_col == -1:
            print("Can't find Icon column")
            return
        # 读取icon_col列内容作为文件名，并设置图片
        # 遍历Icon列，读取文件名并插入图片
        for row in ws.iter_rows(min_row=2, min_col=icon_col + 1, max_col=icon_col + 1):
            for cell in row:
                icon_id = cell.value
                if icon_id:
                    if icon_id == "EM1164_50_0":
                        icon_id = "EM1164_00_0"  # sb capcom
                    icon_path = os.path.join(
                        icon_dir, f"tex_EmIcon_{icon_id}_IMLM4.tex.241106027.png"
                    )
                    if os.path.exists(icon_path):
                        print(f"Found icon {icon_id}")
                        img = Image(icon_path)
                        img.width = 100
                        img.height = 100
                        cell.value = ""
                        # 将图片插入到当前格
                        ws.add_image(
                            img,
                            # openpyxl.utils.get_column_letter(icon_col + 2)
                            openpyxl.utils.get_column_letter(icon_col + 1)
                            + str(cell.row),
                        )
                        # 调整行高
                        ws.row_dimensions[cell.row].height = 80
                    else:
                        print(f"File not found: {icon_path}")

        # 除首行外全部自动换行
        align_wrap = Alignment(wrapText=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = align_wrap

        autofit.style_workbook(writer.book, max_width=60)


if __name__ == "__main__":
    main()
